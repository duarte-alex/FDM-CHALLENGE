import pandas as pd
from collections import defaultdict
import os
from openpyxl import load_workbook
from fastapi import UploadFile
from io import BytesIO


def handle_non_tabular(file: UploadFile) -> bool:
    """
    Parses non-tabular Excel or CSV file with triplet-format rows.

    Args:
        file (UploadFile): Uploaded file object (.xlsx, .xls, .csv)

    Returns:
        bool: True if saving was successful, False otherwise.
    """
    ext = os.path.splitext(file.filename)[-1].lower()
    rows = []

    if ext == ".xlsx":
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        file.file.seek(0)

    elif ext == ".xls":
        df = pd.read_excel(file.file, header=None, engine="xlrd")
        rows = df.values.tolist()
        file.file.seek(0)

    elif ext == ".csv":
        df = pd.read_csv(file.file, header=None)
        rows = df.values.tolist()
        file.file.seek(0)

    else:
        raise ValueError(f"Unsupported file format: {ext}")

    date_row = rows[1]
    dates = list(date_row)[::3]

    date_to_entries = defaultdict(list)

    for row in rows[3:]:
        for i in range(0, len(row), 3):

            start_time = row[i]
            grade = row[i + 1]
            mould_size = row[i + 2]

            if grade is None or str(grade).strip() in ("-", "Grade"):
                continue

            entry = {
                "start_time": start_time,
                "grade": str(grade).strip(),
                "mould_size": mould_size,
            }

            try:
                date_key = dates[i // 3]
                if date_key:
                    date_to_entries[date_key].append(entry)
            except IndexError:
                continue

    # Save output
    os.makedirs("data/processed", exist_ok=True)
    processed_files = []
    for date, entries in date_to_entries.items():
        try:
            df = pd.DataFrame(entries)
            date_str = pd.to_datetime(date).strftime("%Y-%m-%d")
            output_path = f"data/processed/charge_schedule_{date_str}.csv"
            df.to_csv(output_path, index=False)
            processed_files.append(output_path)
        except Exception as e:
            return False

    return True


def sheet_to_pandas(file: UploadFile, skip=1) -> pd.DataFrame:
    """
    Load a spreadsheet into a pandas DataFrame for tabular data.

    Args:
        file (UploadFile): Uploaded file object (.xlsx, .xls, .csv)
        skip (int): Number of rows to skip from the top

    Returns:
        pd.DataFrame: Parsed DataFrame.
    """
    ext = os.path.splitext(file.filename)[-1].lower()

    if ext == ".xlsx":
        return pd.read_excel(
            BytesIO(file.file.read()), engine="openpyxl", skiprows=skip
        )

    elif ext == ".xls":
        return pd.read_excel(BytesIO(file.file.read()), engine="xlrd", skiprows=skip)

    elif ext == ".csv":
        return pd.read_csv(BytesIO(file.file.read()), encoding="utf-8", skiprows=skip)

    else:
        raise ValueError(f"Unsupported file format: {ext}")


def process_steel_grade(
    file: UploadFile, col_name: str = "Quality group"
) -> pd.DataFrame:
    """
    Load a spreadsheet and process with forward fill and dropna operations.

    Args:
        file (UploadFile): Uploaded file object (.xlsx, .xls, .csv)
        col_name (str): Column name to apply forward fill

    Returns:
        pd.DataFrame: Processed DataFrame.
    """
    return (
        sheet_to_pandas(file, skip=1)
        .dropna(axis=1, how="all")
        .assign(**{col_name: lambda df: df[col_name].ffill()})
        .rename(columns={"Grade": "grade_name", "Quality group": "product_group_id"})
        .melt(
            id_vars=["grade_name", "product_group_id"],
            var_name="date",
            value_name="tons",
        )
    )
